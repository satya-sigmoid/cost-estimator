# excel_writer_combined.py
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, NamedStyle
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from decimal import Decimal
import os


def _try_number(v):
    if v is None:
        return ""
    if isinstance(v, (int, float)):
        return v
    if isinstance(v, bool):
        return v
    s = str(v).strip().replace(",", "")
    if s == "":
        return ""
    try:
        if "." in s:
            return float(s)
        return int(s)
    except:
        try:
            return float(Decimal(s))
        except:
            return v


def _normalize_string(v):
    if v is None:
        return ""
    if isinstance(v, list):
        return ", ".join(str(x) for x in v)
    if isinstance(v, dict):
        return "; ".join(f"{k}: {v}" for k, v in v.items())
    return str(v)


def auto_fit_columns(ws, extra_padding=4):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                length = len(str(cell.value)) if cell.value else 0
                max_len = max(max_len, length)
            except:
                pass
        ws.column_dimensions[col_letter].width = max_len + extra_padding


HEADER_FILL = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
HEADER_FONT = Font(bold=True)
TITLE_FONT = Font(size=14, bold=True)

CURRENCY = NamedStyle(name="currency_style")
CURRENCY.number_format = '"$"#,##0.00'

INT_STYLE = NamedStyle(name="int_style")
INT_STYLE.number_format = "#,##0"

NUMBER_STYLE = NamedStyle(name="number_style")
NUMBER_STYLE.number_format = "#,##0.00"

# -------------------------------------------------------------------
# ARCHITECTURE DIAGRAM SHEET (unchanged)
# -------------------------------------------------------------------
def write_architecture_diagram_sheet(wb, image_path, use_case_name):
    ws = wb.create_sheet(f"Architecture_{use_case_name}")

    ws["A1"] = "Architecture Diagram"
    ws["A1"].font = TITLE_FONT

    if not image_path or not os.path.exists(image_path):
        ws["A3"] = "Architecture diagram not available."
        return ws

    img = XLImage(image_path)

    # Resize image if needed (optional but recommended)
    img.width = min(img.width, 900)
    img.height = min(img.height, 600)

    ws.add_image(img, "A3")

    ws.column_dimensions["A"].width = 120
    ws.row_dimensions[3].height = 400

    return ws

# -------------------------------------------------------------------
# BASELINE ASSUMPTION SHEET
# -------------------------------------------------------------------
def write_combined_sheet(wb, baseline_list, cost_components, pipelines):

    ws = wb.create_sheet("Baseline_cost_assumption")

    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws.cell(row=row, column=1).value = "CLOUD COST ESTIMATION OVERVIEW"
    ws.cell(row=row, column=1).font = TITLE_FONT
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
    row += 2

    # ---------------------------------------------------------
    # BASELINE SUMMARY
    # ---------------------------------------------------------
    ws.cell(row=row, column=1).value = "Baseline Summary"
    ws.cell(row=row, column=1).font = HEADER_FONT
    row += 1

    headers = ["Parameter", "Usecase Details", "Source of Assumption", "Notes"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    row += 1

    for entry in baseline_list:
        ws.append([
            entry.get("parameter", ""),
            _normalize_string(entry.get("usecase_details", "")),
            _normalize_string(entry.get("source_of_assumption", "")),
            _normalize_string(entry.get("notes", ""))
        ])
        row += 1

    row += 2  # 2 row gap

    # ---------------------------------------------------------
    # COST COMPONENTS
    # ---------------------------------------------------------
    ws.cell(row=row, column=1).value = "Detailed Cost Components"
    ws.cell(row=row, column=1).font = HEADER_FONT
    row += 1

    headers = ["Cost Component", "Calculation Logic", "Cost ($)", "Source", "Remarks"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    row += 1

    for comp in cost_components:
        cost_val = _try_number(comp.get("cost_usd", ""))
        row_values = [
            comp.get("component", ""),
            comp.get("calculation_logic", ""),
            cost_val,
            comp.get("source", ""),
            comp.get("remarks", "")
        ]
        ws.append(row_values)

        # Cost → currency formatting
        cell = ws.cell(row=row, column=3)
        if isinstance(cost_val, (float, int)):
            cell.number_format = CURRENCY.number_format
        row += 1

    row += 2

    # ---------------------------------------------------------
    # PIPELINE GROUPS
    # ---------------------------------------------------------
    ws.cell(row=row, column=1).value = "Pipeline Groups"
    ws.cell(row=row, column=1).font = HEADER_FONT
    row += 1

    headers = [
        "Pipeline Group", "Data Sources Included", "Refresh Frequency",
        "Runs/Month", "Avg Hours/Run", "Total Hours/Month"
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    row += 1

    for p in pipelines:
        ws.append([
            p.get("pipeline_name", ""),
            _normalize_string(p.get("data_sources_included", "")),
            p.get("refresh_frequency", ""),
            _try_number(p.get("runs_per_month", "")),
            _try_number(p.get("avg_hours_per_run", "")),
            _try_number(p.get("total_hours_per_month", ""))
        ])
        row += 1

    auto_fit_columns(ws)
    return ws

# -------------------------------------------------------------------
# MONTHLY ENVIRONMENT SHEET (unchanged)
# -------------------------------------------------------------------
def write_monthly_environment_sheet(wb, monthly_env, markets=None):
    ws = wb.create_sheet("Yearly_Cost")

    # -----------------------------
    # Prepare market timeline
    # -----------------------------
    market_timeline = [{"market": "M1", "multiplier": 1.0, "start_month": 1}]
    if markets:
        market_timeline.extend(markets)

    active_markets_per_month = {}
    multiplier_per_month = {}

    for m in range(1, 13):
        active = [
            mk["market"]
            for mk in market_timeline
            if mk["start_month"] <= m
        ]
        multiplier_sum = sum(
            mk["multiplier"]
            for mk in market_timeline
            if mk["start_month"] <= m
        )

        active_markets_per_month[m] = "+".join(active)
        multiplier_per_month[m] = multiplier_sum

    # -----------------------------
    # Header rows (shifted right)
    # -----------------------------
    ws.append(["", "Months"] + [f"m{i}" for i in range(1, 13)] + ["Total"])
    ws.append(
        ["", "Markets"] +
        [active_markets_per_month[i] for i in range(1, 13)] +
        [""]
    )

    for row in (1, 2):
        for cell in ws[row]:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            
    # -----------------------------
    # Environment rows (shifted right)
    # -----------------------------
    for env in ("Dev", "QA", "Prod"):
        record = monthly_env.get(env, {})
        base_months = [_try_number(record.get(f"M{i}", 0)) for i in range(1, 13)]

        final_months = []
        for i in range(12):
            val = base_months[i]
            if isinstance(val, (int, float)):
                final_months.append(round(val * multiplier_per_month[i + 1], 2))
            else:
                final_months.append("")

        total = sum(v for v in final_months if isinstance(v, (int, float)))

        ws.append(["", env] + final_months + [total])
    ws["A3"] = "Environment"
    ws["A3"].font = HEADER_FONT
    ws["A3"].fill = HEADER_FILL
    auto_fit_columns(ws)
    return ws


# -------------------------------------------------------------------
# MAIN ENTRY
# -------------------------------------------------------------------
def generate_cost_excel_combined(json_output, file_path, client_name, use_case_name, architecture_image_path=None, markets=None):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    baseline = json_output.get("baseline_summary", [])
    cost_components = json_output.get("detailed_cost_components", [])
    pipelines = json_output.get("pipeline_groups", [])
    monthly_env = json_output.get("monthly_environment_costs", {})
    
    if architecture_image_path:
        write_architecture_diagram_sheet(wb, architecture_image_path, use_case_name)
        
    write_combined_sheet(wb, baseline, cost_components, pipelines)
    # write_monthly_environment_sheet(wb, monthly_env)
    
    # if(markets):
    write_monthly_environment_sheet(wb, monthly_env, markets)

    wb.save(file_path)
    return file_path